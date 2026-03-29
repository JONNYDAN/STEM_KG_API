import argparse
import json
import math
import os
import random
import statistics
import time
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

import requests

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
except Exception as ex:  # pragma: no cover
    raise RuntimeError(
        "openpyxl is required for XLSX export. Please install it (pip install openpyxl)."
    ) from ex


@dataclass
class SeedCase:
    topic: str
    text: str
    relevant_ids: List[str]
    image_keywords: List[str]


SEED_TEXT_CASES: List[SeedCase] = [
    SeedCase("ladybug", "Hãy cho tôi biết vòng đời của con bọ rùa diễn ra như thế nào?", ["105.png"], ["ladybug"]),
    SeedCase("water cycle", "Vòng tuần hoàn nước có các quy trình như thế nào?", ["1483.png"], ["water", "nuoc"]),
    SeedCase("solar system", "Hệ mặt trời là gì vậy?", ["1422.png"], ["solar"]),
    SeedCase("volcano", "Bạn cho tôi một hình ảnh cấu trúc của núi lửa được không?", ["1455.png"], ["volcano"]),
    SeedCase("solar eclipse", "Thông tin về Nhật thực", ["1630.png"], ["eclipse", "solareclipse"]),
    SeedCase("parts of flower", "Các phần của hoa", ["1014.png"], ["hoa", "flower"]),
    SeedCase("electric circuit", "Làm sao để vẽ mạch điện", ["1053.png"], ["mach", "circuit", "so_do"]),
    SeedCase("photosynthesis", "Qúa trình quang hợp diễn ra làm sao?", ["1263.png"], ["photosynthesis", "calvin"]),
]

TEXT_PREFIXES = [
    "Hãy giải thích",
    "Mô tả ngắn gọn",
    "Tôi muốn hiểu",
    "Giúp tôi tìm hiểu",
    "Cho tôi biết",
]

TEXT_SUFFIXES = [
    "bằng sơ đồ dễ hiểu.",
    "theo cách học sinh dễ nắm.",
    "và nêu các ý chính.",
    "kèm hình minh họa nếu có.",
    "một cách trực quan.",
]


def _normalize(value: Any) -> str:
    return str(value or "").strip().lower()


def _unique_preserve_order(items: Sequence[str]) -> List[str]:
    seen = set()
    output: List[str] = []
    for item in items:
        token = _normalize(item)
        if not token or token in seen:
            continue
        seen.add(token)
        output.append(token)
    return output


def _precision_at_k(predicted: List[str], relevant: Set[str], k: int) -> float:
    if k <= 0:
        return 0.0
    top_k = predicted[:k]
    if not top_k:
        return 0.0
    hits = sum(1 for item in top_k if item in relevant)
    return hits / float(k)


def _recall_at_k(predicted: List[str], relevant: Set[str], k: int) -> float:
    if not relevant:
        return 0.0
    hits = sum(1 for item in predicted[:k] if item in relevant)
    return hits / float(len(relevant))


def _f1(precision: float, recall: float) -> float:
    if precision + recall == 0:
        return 0.0
    return 2.0 * precision * recall / (precision + recall)


def _ap_at_k(predicted: List[str], relevant: Set[str], k: int) -> float:
    if not relevant:
        return 0.0
    running_hits = 0
    score = 0.0
    for rank, item in enumerate(predicted[:k], start=1):
        if item in relevant:
            running_hits += 1
            score += running_hits / float(rank)
    return score / float(len(relevant))


def _extract_ranked_diagram_ids(payload: Dict[str, Any]) -> List[str]:
    ranked: List[str] = []

    final_output = payload.get("final_output") or {}
    diagram_id = (final_output.get("diagram") or {}).get("diagram_id")
    if diagram_id:
        ranked.append(str(diagram_id))

    for item in payload.get("query_results") or []:
        results = item.get("results") or {}
        for source_key in ["diagrams", "postgres"]:
            if source_key == "diagrams":
                rows = results.get("diagrams") or []
            else:
                rows = (results.get("postgres") or {}).get("diagrams") or []
            for row in rows:
                d_id = row.get("diagram_id") or row.get("id")
                if d_id:
                    ranked.append(str(d_id))

    return _unique_preserve_order(ranked)


def _collect_image_files(test_images_dir: str) -> List[str]:
    if not os.path.isdir(test_images_dir):
        return []

    allowed_ext = {".png", ".jpg", ".jpeg", ".webp", ".avif", ".bmp"}
    files: List[str] = []
    for name in os.listdir(test_images_dir):
        full_path = os.path.join(test_images_dir, name)
        if not os.path.isfile(full_path):
            continue
        ext = os.path.splitext(name)[1].lower()
        if ext in allowed_ext:
            files.append(full_path)
    return sorted(files)


def _infer_topic_from_image_name(path: str) -> str:
    name = _normalize(os.path.basename(path))
    if "dragonfly" in name:
        return "dragonfly"
    if "hoa" in name or "flower" in name:
        return "parts of flower"
    if "volcano" in name:
        return "volcano"
    if "eclipse" in name:
        return "solar eclipse"
    if "nuoc" in name or "water" in name:
        return "water cycle"
    if "calvin" in name or "photosynthesis" in name:
        return "photosynthesis"
    if "mach" in name or "circuit" in name or "so_do" in name:
        return "electric circuit"
    if "solar" in name:
        return "solar system"
    return "unknown"


def _build_topic_relevance_map() -> Dict[str, List[str]]:
    mapping: Dict[str, List[str]] = {}
    for item in SEED_TEXT_CASES:
        mapping[item.topic] = [str(v) for v in item.relevant_ids]
    return mapping


def _pick_image_for_seed(seed: SeedCase, image_files: List[str]) -> Optional[str]:
    if not image_files:
        return None

    lowered = [p for p in image_files]
    for keyword in seed.image_keywords:
        for path in lowered:
            if keyword in _normalize(os.path.basename(path)):
                return path
    return random.choice(image_files)


def _mutate_text(text: str) -> str:
    if random.random() < 0.35:
        return text

    prefix = random.choice(TEXT_PREFIXES)
    suffix = random.choice(TEXT_SUFFIXES)
    return f"{prefix} {text} {suffix}".strip()


def _post_query(
    base_url: str,
    query_text: Optional[str],
    image_path: Optional[str],
    user_id: str,
    analysis_mode: str,
    timeout: int,
) -> Dict[str, Any]:
    url = f"{base_url.rstrip('/')}/api/integration/query"

    data: Dict[str, str] = {
        "user_id": user_id,
        "analysis_mode": analysis_mode,
    }
    if query_text:
        data["query_text"] = query_text

    files = None
    if image_path and os.path.exists(image_path):
        file_name = os.path.basename(image_path)
        files = {
            "image": (
                file_name,
                open(image_path, "rb"),
                "application/octet-stream",
            )
        }

    try:
        resp = requests.post(url, data=data, files=files, timeout=timeout)
        if not resp.ok:
            raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:300]}")
        payload = resp.json()
        return payload
    finally:
        if files:
            try:
                files["image"][1].close()
            except Exception:
                pass


def _safe_mean(values: List[float]) -> float:
    return statistics.mean(values) if values else 0.0


def _safe_median(values: List[float]) -> float:
    return statistics.median(values) if values else 0.0


def _safe_percentile(values: List[float], pct: float) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    if len(ordered) == 1:
        return ordered[0]
    rank = (len(ordered) - 1) * pct
    lower = math.floor(rank)
    upper = math.ceil(rank)
    if lower == upper:
        return ordered[lower]
    weight = rank - lower
    return ordered[lower] * (1.0 - weight) + ordered[upper] * weight


def _build_conclusion(mode_summary: Dict[str, Dict[str, Any]], top_k: int) -> str:
    lines: List[str] = []
    if not mode_summary:
        return "Không đủ dữ liệu để kết luận."

    evaluable_modes = [
        (mode, stats.get(f"map@{top_k}", 0.0))
        for mode, stats in mode_summary.items()
        if stats.get("evaluated_runs", 0) > 0
    ]

    if evaluable_modes:
        best_mode = max(evaluable_modes, key=lambda item: item[1])[0]
        lines.append(f"Mode có MAP@{top_k} tốt nhất: {best_mode}.")

    fastest_mode = min(mode_summary.items(), key=lambda item: item[1].get("avg_client_total_ms", 0.0))[0]
    lines.append(f"Mode phản hồi nhanh nhất theo trung bình client time: {fastest_mode}.")

    if (
        "text" in mode_summary
        and "both" in mode_summary
        and mode_summary["text"].get("evaluated_runs", 0) > 0
        and mode_summary["both"].get("evaluated_runs", 0) > 0
    ):
        text_map = mode_summary["text"].get(f"map@{top_k}", 0.0)
        both_map = mode_summary["both"].get(f"map@{top_k}", 0.0)
        if both_map >= text_map:
            lines.append("Kết hợp text+image cho độ chính xác truy hồi không kém text-only trên bộ test hiện tại.")
        else:
            lines.append("Text-only đang nhỉnh hơn text+image về MAP trên bộ test hiện tại.")
    elif "text" in mode_summary and "both" in mode_summary:
        lines.append("Chưa đủ ground truth để so sánh trực tiếp chất lượng giữa text-only và text+image.")

    if "image" in mode_summary and mode_summary["image"].get("evaluated_runs", 0) == 0:
        lines.append("Image-only hiện chưa có ground truth đầy đủ nên chủ yếu đánh giá latency/success rate.")

    return "\n".join(lines)


def _write_xlsx(
    output_path: str,
    run_rows: List[Dict[str, Any]],
    mode_summary: Dict[str, Dict[str, Any]],
    overall_summary: Dict[str, Any],
    top_k: int,
    conclusion: str,
) -> None:
    wb = Workbook()

    ws_runs = wb.active
    ws_runs.title = "runs"

    run_headers = [
        "run_id",
        "input_mode",
        "analysis_mode",
        "topic",
        "query_text",
        "image_path",
        "status",
        "error",
        "has_diagram",
        "routing_mode",
        "predicted_top1",
        "predicted_ids",
        "relevant_ids",
        f"precision@{top_k}",
        f"recall@{top_k}",
        f"f1@{top_k}",
        f"ap@{top_k}",
        "client_total_ms",
        "server_total_ms",
        "model_analysis_ms",
        "kg_query_ms",
        "timestamp",
    ]
    ws_runs.append(run_headers)
    for row in run_rows:
        ws_runs.append([
            row.get("run_id"),
            row.get("input_mode"),
            row.get("analysis_mode"),
            row.get("topic"),
            row.get("query_text"),
            row.get("image_path"),
            row.get("status"),
            row.get("error"),
            row.get("has_diagram"),
            row.get("routing_mode"),
            row.get("predicted_top1"),
            ", ".join(row.get("predicted_ids") or []),
            ", ".join(row.get("relevant_ids") or []),
            row.get(f"precision@{top_k}"),
            row.get(f"recall@{top_k}"),
            row.get(f"f1@{top_k}"),
            row.get(f"ap@{top_k}"),
            row.get("client_total_ms"),
            row.get("server_total_ms"),
            row.get("model_analysis_ms"),
            row.get("kg_query_ms"),
            row.get("timestamp"),
        ])

    ws_summary = wb.create_sheet("summary")
    summary_headers = [
        "mode",
        "runs",
        "evaluated_runs",
        "success_rate",
        "diagram_found_rate",
        f"map@{top_k}",
        f"macro_precision@{top_k}",
        f"macro_recall@{top_k}",
        f"macro_f1@{top_k}",
        "avg_client_total_ms",
        "median_client_total_ms",
        "p90_client_total_ms",
        "avg_model_analysis_ms",
        "avg_kg_query_ms",
    ]
    ws_summary.append(summary_headers)

    mode_order = ["text", "image", "both"]
    mode_rows = []
    for mode in mode_order:
        if mode not in mode_summary:
            continue
        stats = mode_summary[mode]
        row = [
            mode,
            stats.get("runs", 0),
            stats.get("evaluated_runs", 0),
            stats.get("success_rate", 0.0),
            stats.get("diagram_found_rate", 0.0),
            stats.get(f"map@{top_k}", 0.0),
            stats.get(f"macro_precision@{top_k}", 0.0),
            stats.get(f"macro_recall@{top_k}", 0.0),
            stats.get(f"macro_f1@{top_k}", 0.0),
            stats.get("avg_client_total_ms", 0.0),
            stats.get("median_client_total_ms", 0.0),
            stats.get("p90_client_total_ms", 0.0),
            stats.get("avg_model_analysis_ms", 0.0),
            stats.get("avg_kg_query_ms", 0.0),
        ]
        ws_summary.append(row)
        mode_rows.append(row)

    ws_overall = wb.create_sheet("overall")
    ws_overall.append(["metric", "value"])
    for key, value in overall_summary.items():
        ws_overall.append([key, value])

    ws_conclusion = wb.create_sheet("conclusion")
    ws_conclusion.append(["Kết luận thử nghiệm"])
    for line in conclusion.split("\n"):
        ws_conclusion.append([line])

    if mode_rows:
        # Chart 1: Average phase latency by mode
        bar = BarChart()
        bar.title = "Average Latency by Input Mode"
        bar.y_axis.title = "Milliseconds"
        bar.x_axis.title = "Mode"

        # Columns J, M, N from summary table (1-based): avg_client_total_ms, avg_model_analysis_ms, avg_kg_query_ms
        data = Reference(ws_summary, min_col=10, max_col=14, min_row=1, max_row=1 + len(mode_rows))
        cats = Reference(ws_summary, min_col=1, min_row=2, max_row=1 + len(mode_rows))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 8
        bar.width = 18
        ws_summary.add_chart(bar, "A8")

        # Chart 2: MAP and F1 by mode
        line = LineChart()
        line.title = f"Quality Metrics by Mode (K={top_k})"
        line.y_axis.title = "Score"
        line.x_axis.title = "Mode"
        data2 = Reference(ws_summary, min_col=6, max_col=9, min_row=1, max_row=1 + len(mode_rows))
        line.add_data(data2, titles_from_data=True)
        line.set_categories(cats)
        line.height = 8
        line.width = 18
        ws_summary.add_chart(line, "A25")

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Large-scale benchmark for STEM retrieval (100-1000+ runs) with text/image/both modes, "
            "phase timing, quality metrics, charts, and XLSX export."
        )
    )
    parser.add_argument("--base-url", default="http://localhost:8000")
    parser.add_argument("--runs", type=int, default=120)
    parser.add_argument("--top-k", type=int, default=5)
    parser.add_argument("--analysis-mode", choices=["basic", "gemini"], default="gemini")
    parser.add_argument("--input-modes", default="text,image,both")
    parser.add_argument("--test-images-dir", default="../test_images")
    parser.add_argument("--user-id", default="benchmark-user")
    parser.add_argument("--timeout", type=int, default=120)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--output-xlsx", default="benchmark_experiment_report.xlsx")
    parser.add_argument("--output-json", default="benchmark_experiment_report.json")
    args = parser.parse_args()

    if args.runs < 1:
        raise ValueError("--runs must be >= 1")

    random.seed(args.seed)

    base_dir = os.path.dirname(os.path.abspath(__file__))
    test_images_dir = args.test_images_dir
    if not os.path.isabs(test_images_dir):
        cwd_candidate = os.path.normpath(os.path.join(os.getcwd(), test_images_dir))
        script_candidate = os.path.normpath(os.path.join(base_dir, test_images_dir))
        if os.path.isdir(cwd_candidate):
            test_images_dir = cwd_candidate
        else:
            test_images_dir = script_candidate

    image_files = _collect_image_files(test_images_dir)
    topic_relevance_map = _build_topic_relevance_map()

    enabled_modes = [m.strip().lower() for m in args.input_modes.split(",") if m.strip()]
    enabled_modes = [m for m in enabled_modes if m in {"text", "image", "both"}]
    if not enabled_modes:
        raise ValueError("No valid --input-modes. Use text,image,both")

    run_rows: List[Dict[str, Any]] = []

    for run_id in range(1, args.runs + 1):
        input_mode = random.choice(enabled_modes)
        seed_case = random.choice(SEED_TEXT_CASES)

        query_text: Optional[str] = None
        image_path: Optional[str] = None
        topic = seed_case.topic
        relevant_ids = [str(v) for v in seed_case.relevant_ids]

        if input_mode == "text":
            query_text = _mutate_text(seed_case.text)
        elif input_mode == "image":
            if image_files:
                image_path = random.choice(image_files)
                inferred_topic = _infer_topic_from_image_name(image_path)
                topic = inferred_topic
                relevant_ids = topic_relevance_map.get(inferred_topic, [])
            else:
                # fallback when image dir is empty
                query_text = _mutate_text(seed_case.text)
                input_mode = "text"
        else:  # both
            query_text = _mutate_text(seed_case.text)
            image_path = _pick_image_for_seed(seed_case, image_files)

        client_started = time.perf_counter()
        status = "ok"
        error = None
        payload: Dict[str, Any] = {}
        try:
            payload = _post_query(
                base_url=args.base_url,
                query_text=query_text,
                image_path=image_path,
                user_id=f"{args.user_id}-{run_id}",
                analysis_mode=args.analysis_mode,
                timeout=args.timeout,
            )
        except Exception as ex:
            status = "error"
            error = str(ex)
        client_total_ms = round((time.perf_counter() - client_started) * 1000.0, 3)

        predicted_ids: List[str] = []
        has_diagram = False
        routing_mode = None
        server_total_ms = None
        model_analysis_ms = None
        kg_query_ms = None

        if status == "ok":
            predicted_ids = _extract_ranked_diagram_ids(payload)
            has_diagram = bool((payload.get("final_output") or {}).get("diagram"))
            routing_mode = (payload.get("query") or {}).get("routing_mode")
            timing = payload.get("timing") or {}
            server_total_ms = timing.get("total_elapsed_ms")
            model_analysis_ms = timing.get("model_analysis_ms")
            kg_query_ms = timing.get("kg_query_ms")

        relevant_set = set(_unique_preserve_order(relevant_ids))
        eval_enabled = status == "ok" and bool(relevant_set)

        precision = _precision_at_k(predicted_ids, relevant_set, args.top_k) if eval_enabled else None
        recall = _recall_at_k(predicted_ids, relevant_set, args.top_k) if eval_enabled else None
        f1 = _f1(precision, recall) if eval_enabled and precision is not None and recall is not None else None
        ap = _ap_at_k(predicted_ids, relevant_set, args.top_k) if eval_enabled else None

        run_rows.append(
            {
                "run_id": run_id,
                "input_mode": input_mode,
                "analysis_mode": args.analysis_mode,
                "topic": topic,
                "query_text": query_text,
                "image_path": image_path,
                "status": status,
                "error": error,
                "has_diagram": has_diagram,
                "routing_mode": routing_mode,
                "predicted_top1": predicted_ids[0] if predicted_ids else None,
                "predicted_ids": predicted_ids,
                "relevant_ids": sorted(list(relevant_set)),
                f"precision@{args.top_k}": round(precision, 6) if precision is not None else None,
                f"recall@{args.top_k}": round(recall, 6) if recall is not None else None,
                f"f1@{args.top_k}": round(f1, 6) if f1 is not None else None,
                f"ap@{args.top_k}": round(ap, 6) if ap is not None else None,
                "client_total_ms": client_total_ms,
                "server_total_ms": server_total_ms,
                "model_analysis_ms": model_analysis_ms,
                "kg_query_ms": kg_query_ms,
                "timestamp": datetime.now(timezone.utc).isoformat(),
            }
        )

    mode_summary: Dict[str, Dict[str, Any]] = {}
    for mode in sorted({row["input_mode"] for row in run_rows}):
        subset = [row for row in run_rows if row["input_mode"] == mode]
        eval_subset = [row for row in subset if row[f"ap@{args.top_k}"] is not None]

        success_rate = sum(1 for row in subset if row["status"] == "ok") / float(len(subset)) if subset else 0.0
        diagram_rate = sum(1 for row in subset if row["has_diagram"]) / float(len(subset)) if subset else 0.0

        mode_summary[mode] = {
            "runs": len(subset),
            "evaluated_runs": len(eval_subset),
            "success_rate": round(success_rate, 6),
            "diagram_found_rate": round(diagram_rate, 6),
            f"map@{args.top_k}": round(_safe_mean([row[f"ap@{args.top_k}"] for row in eval_subset]), 6),
            f"macro_precision@{args.top_k}": round(_safe_mean([row[f"precision@{args.top_k}"] for row in eval_subset]), 6),
            f"macro_recall@{args.top_k}": round(_safe_mean([row[f"recall@{args.top_k}"] for row in eval_subset]), 6),
            f"macro_f1@{args.top_k}": round(_safe_mean([row[f"f1@{args.top_k}"] for row in eval_subset]), 6),
            "avg_client_total_ms": round(_safe_mean([row["client_total_ms"] for row in subset]), 3),
            "median_client_total_ms": round(_safe_median([row["client_total_ms"] for row in subset]), 3),
            "p90_client_total_ms": round(_safe_percentile([row["client_total_ms"] for row in subset], 0.9), 3),
            "avg_model_analysis_ms": round(_safe_mean([row["model_analysis_ms"] for row in subset if row["model_analysis_ms"] is not None]), 3),
            "avg_kg_query_ms": round(_safe_mean([row["kg_query_ms"] for row in subset if row["kg_query_ms"] is not None]), 3),
        }

    overall_eval = [row for row in run_rows if row[f"ap@{args.top_k}"] is not None]
    overall_summary = {
        "runs": len(run_rows),
        "evaluated_runs": len(overall_eval),
        "success_rate": round(sum(1 for row in run_rows if row["status"] == "ok") / float(len(run_rows)), 6),
        "diagram_found_rate": round(sum(1 for row in run_rows if row["has_diagram"]) / float(len(run_rows)), 6),
        f"map@{args.top_k}": round(_safe_mean([row[f"ap@{args.top_k}"] for row in overall_eval]), 6),
        f"macro_precision@{args.top_k}": round(_safe_mean([row[f"precision@{args.top_k}"] for row in overall_eval]), 6),
        f"macro_recall@{args.top_k}": round(_safe_mean([row[f"recall@{args.top_k}"] for row in overall_eval]), 6),
        f"macro_f1@{args.top_k}": round(_safe_mean([row[f"f1@{args.top_k}"] for row in overall_eval]), 6),
        "avg_client_total_ms": round(_safe_mean([row["client_total_ms"] for row in run_rows]), 3),
        "avg_model_analysis_ms": round(_safe_mean([row["model_analysis_ms"] for row in run_rows if row["model_analysis_ms"] is not None]), 3),
        "avg_kg_query_ms": round(_safe_mean([row["kg_query_ms"] for row in run_rows if row["kg_query_ms"] is not None]), 3),
    }

    conclusion = _build_conclusion(mode_summary, args.top_k)

    report = {
        "config": {
            "base_url": args.base_url,
            "runs": args.runs,
            "top_k": args.top_k,
            "analysis_mode": args.analysis_mode,
            "input_modes": enabled_modes,
            "test_images_dir": test_images_dir,
            "seed": args.seed,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        },
        "overall_summary": overall_summary,
        "mode_summary": mode_summary,
        "conclusion": conclusion,
        "runs": run_rows,
    }

    with open(args.output_json, "w", encoding="utf-8") as fp:
        json.dump(report, fp, ensure_ascii=False, indent=2)

    _write_xlsx(
        output_path=args.output_xlsx,
        run_rows=run_rows,
        mode_summary=mode_summary,
        overall_summary=overall_summary,
        top_k=args.top_k,
        conclusion=conclusion,
    )

    print("=" * 90)
    print("Benchmark Summary")
    print("=" * 90)
    for key, value in overall_summary.items():
        print(f"{key}: {value}")
    print("-" * 90)
    print("Conclusion:")
    print(conclusion)
    print("=" * 90)
    print(f"Saved JSON: {args.output_json}")
    print(f"Saved XLSX: {args.output_xlsx}")


if __name__ == "__main__":
    main()
